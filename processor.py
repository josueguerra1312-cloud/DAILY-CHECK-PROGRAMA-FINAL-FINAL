from __future__ import annotations
import copy, io, re
from datetime import datetime, time
from pathlib import Path
from openpyxl import load_workbook


def txt(v):
    if v is None: return ""
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v).strip()


def ac(v): return re.sub(r"\s+", "", txt(v).upper())


def seconds(v):
    if isinstance(v, datetime): v = v.time()
    if isinstance(v, time): return v.hour*3600+v.minute*60+v.second
    if isinstance(v, (int,float)): return int((float(v)%1)*86400)
    for f in ("%H:%M:%S","%H:%M","%I:%M %p"):
        try:
            d=datetime.strptime(txt(v),f); return d.hour*3600+d.minute*60+d.second
        except ValueError: pass
    return None


def elapsed(a,d):
    a,d=seconds(a),seconds(d)
    if a is None or d is None: return -1
    return d-a if d>=a else 86400-a+d


def copy_style(ws, source, target):
    ws.row_dimensions[target].height=ws.row_dimensions[source].height
    for c in range(1,ws.max_column+1):
        s,t=ws.cell(source,c),ws.cell(target,c)
        if s.has_style: t._style=copy.copy(s._style)
        t.number_format=s.number_format
        t.alignment=copy.copy(s.alignment)
        t.protection=copy.copy(s.protection)


def program_groups(source):
    wb=load_workbook(source,data_only=False)
    ws=wb["Programa"] if "Programa" in wb.sheetnames else wb.active
    h={txt(ws.cell(1,c).value).upper():c for c in range(1,ws.max_column+1)}
    required={"AC","WO","TASK","DESCRIPTION"}
    if not required.issubset(h): raise ValueError("El programa requiere AC, WO, TASK y DESCRIPTION en la fila 1.")
    out=[]; current=None; ca=cw=""
    for r in range(2,ws.max_row+1):
        ca=ac(ws.cell(r,h["AC"]).value) or ca
        cw=txt(ws.cell(r,h["WO"]).value) or cw
        task=txt(ws.cell(r,h["TASK"]).value); desc=txt(ws.cell(r,h["DESCRIPTION"]).value)
        if not task and not desc: continue
        if current is None or current["ac"]!=ca or current["wo"]!=cw:
            current={"ac":ca,"wo":cw,"tasks":[]}; out.append(current)
        current["tasks"].append((task,desc))
    return [g for g in out if g["ac"] and g["tasks"]]


def find_limits(ws):
    header=None
    for r in range(1,min(20,ws.max_row)+1):
        if [txt(ws.cell(r,c).value).upper() for c in range(1,5)]==["A/C","FL","ARR","DEPT"]: header=r; break
    if header is None: raise ValueError("No se encontro A/C, FL, ARR, DEPT en GDL.")
    ron=next((r for r in range(header+1,ws.max_row+1) if txt(ws.cell(r,1).value).upper()=="RON AC"),None)
    if ron is None: raise ValueError("No se encontro RON AC.")
    start=header+1
    while start<ron and txt(ws.cell(start,2).value).upper() in {"RTO","STORAGE",""}: start+=1
    return header,start,ron-1


def add_storage(ws, groups):
    used=set(); r=3
    while r<=ws.max_row:
        aircraft=ac(ws.cell(r,1).value)
        if txt(ws.cell(r,2).value).upper()!="STORAGE": r+=1; continue
        matches=[(i,g) for i,g in enumerate(groups) if g["ac"]==aircraft]
        if not matches: r+=1; continue
        end=r
        while end+1<=ws.max_row and not ac(ws.cell(end+1,1).value): end+=1
        additions=[(i,t) for i,g in matches for t in g["tasks"]]
        n=len(additions); at=end+1
        merges=list(ws.merged_cells.ranges)
        for rg in merges: ws.unmerge_cells(str(rg))
        ws.insert_rows(at,n)
        for rg in merges:
            shift=n if rg.min_row>=at else 0
            grow=n if rg.min_row<=r<=rg.max_row or rg.min_row<=end<=rg.max_row else 0
            ws.merge_cells(start_row=rg.min_row+shift,start_column=rg.min_col,end_row=rg.max_row+shift+grow,end_column=rg.max_col)
        for off,(i,(task,desc)) in enumerate(additions):
            copy_style(ws,end,at+off); ws.cell(at+off,6).value=task; ws.cell(at+off,7).value=desc; used.add(i)
        r=at+n
    return used


def generate_combined(program_source, template_source, output_path=None):
    groups=program_groups(program_source)
    wb=load_workbook(template_source,data_only=False)
    if "GDL" not in wb.sheetnames: raise ValueError("La plantilla requiere la hoja GDL.")
    ws=wb["GDL"]
    used=add_storage(ws,groups); groups=[g for i,g in enumerate(groups) if i not in used]
    _,start,end=find_limits(ws)
    flights=[]
    for r in range(start,end+1):
        if ac(ws.cell(r,1).value): flights.append({"row":r,"ac":ac(ws.cell(r,1).value),"fl":ws.cell(r,2).value,"arr":ws.cell(r,3).value,"dept":ws.cell(r,4).value})
    assigned={}
    for g in groups:
        ids=[i for i,f in enumerate(flights) if f["ac"]==g["ac"]]
        if ids: assigned.setdefault(max(ids,key=lambda i:elapsed(flights[i]["arr"],flights[i]["dept"])),[]).append(g)
    rows=[]
    for i,f in enumerate(flights):
        gs=assigned.get(i,[])
        if gs:
            tasks=[(g,t) for g in gs for t in g["tasks"]]
            rows.append((f,gs,tasks))
        else: rows.append((f,[],[]))
    old_summary=end+1; summary=[]
    for r in range(old_summary,ws.max_row+1):
        summary.append([(ws.cell(r,c).value,copy.copy(ws.cell(r,c)._style)) for c in range(1,ws.max_column+1)])
    total=sum(len(tasks) if tasks else 1 for _,_,tasks in rows)
    for rg in list(ws.merged_cells.ranges):
        if rg.max_row>=start: ws.unmerge_cells(str(rg))
    if total>(end-start+1): ws.insert_rows(old_summary,total-(end-start+1))
    for r in range(start,start+total):
        for c in range(1,ws.max_column+1): ws.cell(r,c).value=None
        copy_style(ws,start,r)
    out=start
    for f,gs,tasks in rows:
        if not tasks:
            for c,v in enumerate((f["ac"],f["fl"],f["arr"],f["dept"]),1): ws.cell(out,c).value=v
            a,d=seconds(f["arr"]),seconds(f["dept"])
            ws.cell(out,5).value="TRANSIT CHECK / RON" if a is not None and d is not None and d<=a else "TRANSIT CHECK"
            ws.merge_cells(start_row=out,start_column=5,end_row=out,end_column=8); out+=1; continue
        first=out; last=out+len(tasks)-1
        for c,v in enumerate((f["ac"],f["fl"],f["arr"],f["dept"]),1): ws.cell(first,c).value=v
        ws.cell(first,5).value="\n".join(dict.fromkeys(g["wo"] for g in gs if g["wo"]))
        for c in range(1,6):
            if last>first: ws.merge_cells(start_row=first,start_column=c,end_row=last,end_column=c)
        for rr,(_,task) in enumerate(tasks,first): ws.cell(rr,6).value=task[0]; ws.cell(rr,7).value=task[1]
        out=last+1
    new_summary=start+total
    for off,row in enumerate(summary):
        for c,(v,style) in enumerate(row,1): ws.cell(new_summary+off,c).value=v; ws.cell(new_summary+off,c)._style=style
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode="auto"
    buffer=io.BytesIO(); wb.save(buffer); data=buffer.getvalue()
    if output_path: Path(output_path).write_bytes(data)
    return data
